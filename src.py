import os.path
from re import search
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, DEFAULT_FONT, numbers
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter

import tkinter as tk

month_name = ['มกราคม',
              'กุมภาพันธ์',
              'มีนาคม',
              'เมษายน',
              'พฤษภาคม',
              'มิถุนายน',
              'กรกฎาคม',
              'สิงหาคม',
              'กันยายน',
              'ตุลาคม',
              'พฤศจิกายน',
              'ธันวาคม']

thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))
double_border = Border(left=Side(style='thin'), 
                       right=Side(style='thin'), 
                       top=Side(style='thin'), 
                       bottom=Side(style='double'))

program_code = {'DSI': 12240500161}

def generate_report(input_files, output_path, prog_code=program_code['DSI']):
    wb = Workbook()
    wb.remove(wb.active)

    dates = []
    total_stats = []
    dfs = []
    for excel_path in input_files:
        excel_fn = os.path.basename(excel_path)
        # extract transactions' date
        match = search(r'\d{4}\d{2}\d{2}', excel_fn)
        date = datetime.strptime(match.group(), '%Y%m%d').date()
        dates.append(date)
        # extract transactions' summary
        df = extract_excel(excel_path, prog_code)
        # export to excel
        total_stat = write_oneday_report(df, wb, date)
        total_stats.extend(total_stat)
        dfs.append(df)

    # warning when data of multiple months are provided
    month = list(set([(d.month, d.year) for d in dates]))
    if len(month) > 1:
        tk.messagebox.showerror(title="Error", message="Files contain data from different months")
    
    write_summary_report(wb, dfs)

    write_overall_report(wb, total_stats, month)

    _font = Font(name="TH SarabunPSK", sz=16)
    {k: setattr(DEFAULT_FONT, k, v) for k, v in _font.__dict__.items()}
    wb.save(output_path)
    return dfs

def extract_excel(path, prog_code):
    cols = ["รายได้คณะ", "กองทุนคณะ.1", "รายได้คณะ.2", "รวม"]
    
    df = pd.read_excel(path, header=1, sheet_name="Detail", dtype={'เลขทะเบียน':str})
    df[cols] = df[cols].astype('float64')   # convert all numbers to float
    df = df.assign(dep_income=df['รายได้คณะ'] + df['กองทุนคณะ.1'] + df['รายได้คณะ.2'])

    df = df[df['รหัสหลักสูตร'] == prog_code]  # keep only only DSI students
    df_filtered = df[['วิทยาเขต', 
                      'คณะ', 
                      'เลขทะเบียน', 
                      'เลขที่ใบเสร็จ', 
                      'ชื่อ นามสกุล', 
                      'ปีการศึกษา', 
                      'ภาค', 
                      'รายได้คณะ', 
                      'กองทุนคณะ.1', 
                      'รายได้คณะ.2', 
                      'dep_income', 
                      'รวม']].copy()
    df_filtered['รหัส'] = df_filtered.apply(lambda row: row['เลขทะเบียน'].strip()[:2], axis = 1)
    return df_filtered.set_index(['ปีการศึกษา', 'ภาค', 'รหัส'])


def write_summary_report(wb, dfs):
    df = pd.concat(dfs)
    df = df[['รายได้คณะ', 'กองทุนคณะ.1', 'รายได้คณะ.2', 'dep_income', 'รวม']]

    for academic_term, d in df.groupby(level=[0,1]):
        # calculate summary
        ov = d.groupby(level=2).agg(sum)
        ov['count'] = d.groupby(level=2).size()
        ov = ov[['count', 'รวม', 'รายได้คณะ', 'กองทุนคณะ.1', 'รายได้คณะ.2', 'dep_income']]

        # write to excel 
        ws = wb.create_sheet("overview_{}_{}".format(*academic_term))
        
        widths = [15, 15, 15, 15, 15, 15, 15]
        for i, w in zip(range(1,8), widths):
            c = get_column_letter(i)
            ws.column_dimensions[c].width = w

        ws.append([None, None, None, None, None, None, None])
        ws.append([None, None, None, 'ค่าหน่วยกิต', 'ค่าธรรมเนียมพิเศษ', 'ค่าธรรมเนียมเพื่อการศึกษาฯ', 'รวมเงินโอนเข้าคณะ'])

        ws.merge_cells('D1:G1')
        cell = ws.cell(row=1, column=4)
        cell.value = 'รายได้คณะ'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
                
        for i, t in enumerate(['รหัส', 'จำนวน น.ศ. (คน)', 'ค่าลงทะเบียน (บาท)']):
            row = chr(i + ord('A'))
            ws.merge_cells(f'{row}1:{row}2')
            cell = ws.cell(row=1, column=i+1)
            cell.value = t
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, row in ov.iterrows():
            ws.append([i] + row.tolist())
        ws.append(['รวมทั้งสิ้น'] + ov.sum().tolist())

        for row in ws[f'C3:G{3+len(ov)}']:
            for cell in row:
                cell.number_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'

        for row in ws[f"D2:G2"]:
            for cell in row:
                cell.font = Font(bold=True)
                
        for row in ws[f"A{3+len(ov)}:G{3+len(ov)}"]:
            for cell in row:
                cell.font = Font(bold=True)

        for row in ws[f'A1:G{3+len(ov)}']:
            for cell in row:
                cell.border = thin_border

def write_overall_report(wb, total_stats, month):
    # group data together
    data = {}
    for k, v in total_stats:
        data[k] = data.setdefault(k, []) + [v]

    for k, v in data.items():
        ws = wb.create_sheet(month_name[month[0][0]-1] + f"_{k[0]}_{k[1]}")

        widths = [50, 15, 15, 15, 15, 15]
        for i, w in zip(range(1,8), widths):
            c = get_column_letter(i)
            ws.column_dimensions[c].width = w
        ws.row_dimensions[3].height = 90
        ws.row_dimensions[4].height = 60

        year = k[0]
        sem = k[1]

        ws.merge_cells('A1:F1')
        cell = ws.cell(row=1, column=1)
        cell.value = f'รายรับค่าจดทะเบียน {sem}/{year} ประจำเดือน {month_name[month[0][0]-1]} {month[0][1]+543}'

        ws.merge_cells('A2:F2')  
        cell = ws.cell(row=2, column=1)
        cell.value = 'หลักสูตรวิทยาศาสตรบัณฑิต สาขาวิชาวิทยาศาสตร์และนวัตกรรมข้อมูล'

        ws.append([None, '18\r\nค่าหน่วยกิต', '01\nค่าธรรมเนียมเพื่อการศึกษาและพัฒนามหาวิทยาลัย', '31 / 98\nค่าธรรมเนียมพิเศษ (อื่นๆ)', None, None])
        ws.append([None, 'รายได้คณะ', 'กองทุนคณะ', 'รายได้คณะ', 'รวม\nรายได้คณะ', 'รวม\nค่าลงทะเบียน'])
        ws.merge_cells('A3:A4')
        cell = ws.cell(row=3, column=1)
        cell.value = 'รวมจำนวนนักศึกษาของแต่ละวัน'
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

        for cell in ws["3:3"]:
                cell.border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'))
            
        for cell in ws["4:4"]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        bottom=Side(style='thin'))

        for cell in ws['B3:B4']:
            cell[0].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell[0].fill = PatternFill(patternType='solid', fgColor="FFFF00")
            cell[0].border = thin_border
            cell[0].font = Font(bold=True)
        for cell in ws['C3:C4']:
            cell[0].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell[0].fill = PatternFill(patternType='solid', fgColor="E2EFDA")
            cell[0].border = thin_border
            cell[0].font = Font(bold=True)
        for cell in ws['D3:D4']:
            cell[0].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell[0].fill = PatternFill(patternType='solid', fgColor="FFF2CC")
            cell[0].border = thin_border
            cell[0].font = Font(bold=True)
        for cell in ws['E3:E4']:
            cell[0].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell[0].fill = PatternFill(patternType='solid', fgColor="92D050")
        for cell in ws['F3:F4']:
            cell[0].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell[0].fill = PatternFill(patternType='solid', fgColor="8EA9DB")
            cell[0].font = Font(color="FFFFFF", bold=True)
        
        n = len(v)+4
        for total_list, total, date, res in v:
            if len(total_list) == 1:
                total_str = total_list[0]
                # row_no += 1
            else:
                total_detail = ", ".join(total_list[:-1]) + " และ " + total_list[-1] + "\n"
                # ws.append([total_detail] + [None]*6)
                total_str = total_detail + f"(รวมนศ.ทั้งหมด จำนวน {total} คน)"
                # row_no += 2

            ws.append([total_str + "              วันที่ " + date.strftime("%d/%#m/") + str(date.year+543)] + res)
        ws.append(["รวม", f"=SUM(B5:B{n})", f"=SUM(C5:C{n})", f"=SUM(D5:D{n})", f"=SUM(E5:E{n})", f"=SUM(F5:F{n})"])
        for row in ws[f'A5:A{n+1}']:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                cell.border = thin_border
        for row in ws[f'B5:F{n+1}']:
            for cell in row:
                cell.number_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
                cell.border = thin_border
        for row in ws[f'A{n+1}:F{n+1}']:
            for cell in row:
                cell.font = Font(bold=True)
                cell.border = thin_border


def write_oneday_report(df, wb, date):
    total_stat = []
    for d, new_df in df.groupby(level=[0,1]):
        ws = wb.create_sheet(date.strftime("%d%m%Y"))
        
        # set header
        widths = [8, 12, 10, 10, 18, 7, 5, 10, 10, 10, 10, 10]
        for i, w in zip(range(1,13), widths):
            c = get_column_letter(i)
            ws.column_dimensions[c].width = w
        ws.row_dimensions[3].height = 90
        ws.row_dimensions[4].height = 60
        
        ws.merge_cells('A1:L1')
        cell = ws.cell(row=1, column=1)
        cell.value = f'รายรับค่าจดทะเบียน {d[1]}/{d[0]} ณ วันที่ {date.day} {month_name[date.month-1]} {date.year+543}'
        
        ws.merge_cells('A2:L2')  
        cell = ws.cell(row=2, column=1)
        cell.value = 'หลักสูตรวิทยาศาสตรบัณฑิต สาขาวิชาวิทยาศาสตร์และนวัตกรรมข้อมูล'
        
        ws.append([None, None, None, None, None, None, None, '18\r\nค่าหน่วยกิต', '01\nค่าธรรมเนียมเพื่อการศึกษาและพัฒนามหาวิทยาลัย', '31 / 98\nค่าธรรมเนียมพิเศษ (อื่นๆ)', None, None])
        ws.append(['วิทยาเขต', 'คณะ', 'เลขทะเบียน', 'เลขที่ใบเสร็จ', 'ชื่อ นามสกุล', 'ปีการศึกษา', 'ภาค', 'รายได้คณะ', 'กองทุนคณะ', 'รายได้คณะ', 'รวม\nรายได้คณะ', 'รวม\nค่าลงทะเบียน'])
        
        for cell in ws["3:3"]:
            cell.border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'))
        
        for cell in ws["4:4"]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        bottom=Side(style='thin'))
        
        for cell in ws['H3:H4']:
            cell[0].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell[0].fill = PatternFill(patternType='solid', fgColor="FFFF00")
            cell[0].border = thin_border
            cell[0].font = Font(bold=True)
        for cell in ws['I3:I4']:
            cell[0].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell[0].fill = PatternFill(patternType='solid', fgColor="E2EFDA")
            cell[0].border = thin_border
            cell[0].font = Font(bold=True)
        for cell in ws['J3:J4']:
            cell[0].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell[0].fill = PatternFill(patternType='solid', fgColor="FFF2CC")
            cell[0].border = thin_border
            cell[0].font = Font(bold=True)
        for cell in ws['K3:K4']:
            cell[0].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell[0].fill = PatternFill(patternType='solid', fgColor="92D050")
        for cell in ws['L3:L4']:
            cell[0].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            cell[0].fill = PatternFill(patternType='solid', fgColor="8EA9DB")
            cell[0].font = Font(color="FFFFFF", bold=True)

        # write rows
        row_no = 4
        for ind, df_code in new_df.groupby(level=2):
            for i, r in enumerate(dataframe_to_rows(df_code, index=False, header=False)):
                ws.append(r[:5] + [*d] + r[5:])
                for cell in ws[f"{row_no+i+1}:{row_no+i+1}"]:
                    cell.font = Font(bold=False)
                    cell.border = thin_border
                for cell in ws[f"H{row_no+i+1}:L{row_no+i+1}"][0]:
                    cell.alignment = Alignment(horizontal="right")
            
            res = df_code.agg('sum').tolist()
            ws.append([None]*7 + res[5:])
            
            row_no = row_no + len(df_code) + 1
            ws.merge_cells(f'A{row_no}:G{row_no}')  
            cell = ws.cell(row=row_no, column=1)
            cell.value = f'รหัส {ind} จำนวน {len(df_code)} คน'
            cell.alignment = Alignment(horizontal="center")
            
            for cell in ws[f"{row_no}:{row_no}"]:
                cell.font = Font(bold=True)
                cell.border = double_border
            for cell in ws[f"H{row_no}:L{row_no}"][0]:
                cell.alignment = Alignment(horizontal="right")

        res = new_df.agg('sum').tolist()
        ws.append([None]*7 + res[5:])
        
        row_no = row_no + 1
        ws.merge_cells(f'A{row_no}:G{row_no}')
        cell = ws.cell(row=row_no, column=1)
        
        code_list = new_df.index.unique(level=2)
        if len(code_list) > 1:
            code_str = ", ".join(code_list[:-1]) + ' และ ' + code_list[-1]
        else:
            code_str = code_list[0]
            
        cell.value = f'รวม {code_str} จำนวน {len(new_df)} คน'
        cell.alignment = Alignment(horizontal="center")
        
        for cell in ws[f"{row_no}:{row_no}"]:
            cell.font = Font(bold=True)
            cell.border = double_border
        
        for row in ws[f'H5:L{row_no}']:
            for cell in row:
                cell.number_format = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
        
        # get total stats for overall report
        total_list = []
        total = 0
        for ind, df_code in new_df.groupby(level=2):
            total_list.append(f'ยอดรวมรหัส {ind} จำนวนนศ. {len(df_code)} คน')
            total += len(df_code)
        total_stat.append((d, [total_list, total, date, res[5:]]))
    return total_stat